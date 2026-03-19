"""
Core PDGM lookup logic.

This module provides the CSV-driven PDGM lookup, Claude AI mapping,
follow-up question generation, and documentation/assessment helpers.
It is imported by blueprints — it does NOT create a Flask app.
"""

import os
import csv
import re
import json
from pathlib import Path
from datetime import datetime, timezone
from difflib import SequenceMatcher

import anthropic

from services.reimbursement_service import (
    ReimbursementService,
    ai_map_phrase_to_code_with_payment,
    extract_pdgm_code_from_response,
)
from enhanced_pdgm_functions import (
    get_candidate_codes_enhanced,
    get_fallback_followup,
    validate_pdgm_response,
    expand_with_medical_synonyms,
)
import enhanced_prompt_manager as prompt_manager

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
CLAUDE_MODEL = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-20250514")
AI_LOG_PATH = Path(os.getenv("AI_LOG_PATH", "ai_logs.jsonl"))

_root = Path(__file__).resolve().parent
CSV_PATH = os.getenv("PDGM_CSV_PATH", str(_root / "CCFTF_pdgm_diagnosis_mapping.csv"))
EXCLUDED_XLSX_PATH = os.getenv(
    "EXCLUDED_XLSX_PATH",
    str(_root / "section111_excluded_icd10_fy2026.xlsx"),
)

# ---------------------------------------------------------------------------
# ICD-10 helpers (also in services/pdgm/impl.py — kept here for direct use)
# ---------------------------------------------------------------------------

def normalize_icd10(code: str) -> str:
    if not isinstance(code, str):
        return ""
    return code.replace('.', '').replace(' ', '').upper()


def normalize_query(q: str) -> str:
    if not isinstance(q, str):
        return ""
    return q.lower().strip()


def load_icd10_map(csv_path=None):
    code_map = {}
    path = csv_path or CSV_PATH
    with open(path, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            code_map[normalize_icd10(row['icd10_code'])] = row
    return code_map


def load_excluded_codes():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCLUDED_XLSX_PATH, read_only=True, data_only=True)
        excluded = set()
        for ws in wb.worksheets:
            header = [str(c.value).strip().upper() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            code_idx = next((i for i, h in enumerate(header) if h == 'CODE'), 0)
            for row in ws.iter_rows(min_row=2):
                val = row[code_idx].value
                if isinstance(val, str) and val.strip():
                    excluded.add(normalize_icd10(val))
        wb.close()
        return excluded
    except Exception:
        return set()


# Load data at import time
icd10_data = load_icd10_map()
excluded_codes = load_excluded_codes()


def format_icd10(code: str) -> str:
    if len(code) > 3 and code[3] != '.':
        return code[:3] + '.' + code[3:]
    return code


def format_pdgm_group(name: str) -> str:
    if not name:
        return ''
    name = name.upper()
    if name.startswith('MMTA_'):
        rest = name[5:].replace('GI_GU', 'GI/GU').replace('_', ' ').title()
        return f"MMTA - {rest}"
    if name.endswith('_REHAB'):
        return name.replace('_', ' ').title()
    if name == 'BEHAVE_HEALTH':
        return 'Behavioral Health'
    if name == 'NO_GROUP':
        return 'No Group'
    return name.replace('_', ' ').title()


def search_icd10_codes(prefix: str, limit: int = 10):
    norm = normalize_icd10(prefix)
    results = []
    for code in sorted(icd10_data):
        if code.startswith(norm):
            row = icd10_data[code]
            results.append({"code": format_icd10(row["icd10_code"]), "description": row.get("description", "")})
            if len(results) >= limit:
                break
    return results


def get_candidate_codes(phrase, limit=15):
    return get_candidate_codes_enhanced(phrase, icd10_data, limit)


def extract_icd10_codes(text: str):
    pattern = r"\b[A-TV-Z][0-9][0-9A-Z]{1,6}(?:\.[0-9A-Z]{1,4})?\b"
    return [normalize_icd10(c) for c in re.findall(pattern, text.upper())]


def validate_ai_output(text: str):
    codes = extract_icd10_codes(text)
    mismatches = [c for c in codes if c not in icd10_data]
    return len(mismatches) == 0, mismatches


def parse_pdgm_details(text):
    """Extract core PDGM details from a response string or dict."""
    if isinstance(text, dict):
        return text
    details = {}
    for line in str(text).split("\n"):
        line = line.strip()
        low = line.lower()
        if low.startswith("focus of care:"):
            details["focus_of_care"] = line.split(":", 1)[1].strip()
        elif low.startswith("description:"):
            details["description"] = line.split(":", 1)[1].strip()
        elif low.startswith("comorbidity group:"):
            details["comorbidity_group"] = line.split(":", 1)[1].strip()
        elif low.startswith("billable:"):
            details["billable"] = line.split(":", 1)[1].strip()
        if len(details) >= 4:
            break
    return details


def log_ai_interaction(model: str, query: str, response: str, mismatches=None):
    entry = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "model": model,
        "query": query,
        "response": response,
        "mismatches": mismatches or [],
    }
    try:
        with open(AI_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception:
        pass


def is_valid_ai_response(response):
    if not response or not isinstance(response, str):
        return False
    error_indicators = [
        "none of the provided codes", "no valid pdgm", "cannot be selected",
        "no appropriate code", "openai error", "rate limit exceeded",
        "error:", "failed to", "unable to",
    ]
    if any(ind in response.lower() for ind in error_indicators):
        return False
    if not re.search(r"\b[A-TV-Z][0-9][0-9A-Z]{1,6}(?:\.[0-9A-Z]{1,4})?\b", response):
        return False
    required = ["focus of care:", "description:", "pdgm group:"]
    if not any(el in response.lower() for el in required):
        return False
    return True

# ---------------------------------------------------------------------------
# Anthropic client + AI functions
# ---------------------------------------------------------------------------
client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None


def _call_claude(system_prompt: str, user_prompt: str, max_tokens: int = 350, temperature: float = 0.1) -> str:
    """Helper to call Claude API."""
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=max_tokens,
        system=system_prompt,
        messages=[{"role": "user", "content": user_prompt}],
        temperature=temperature,
    )
    return resp.content[0].text.strip()


def ai_map_phrase_to_code(phrase):
    if not client:
        return "Anthropic API key not configured."
    system_prompt = (
        "You are a Utilization Review and Coding Compliance Nurse with 20+ years of Home Health experience, certified in HCS-D, HCS-O, CCS, and Medicare billing. "
        "Given a clinical phrase, select only PDGM-valid ICD-10 code(s) from the provided list, recognizing synonymous phrasing, "
        "prioritizing compliant primary-diagnosis selection, and append one line beginning with \"Reason:\" that briefly states why each code was chosen. "
        "Respond for each match using this format:\n"
        "Focus of Care: <ICD10 no dot> <Description> <PDGM Code> <PDGM Group>  ICD-10: <ICD10 with dot>\n"
        "Description: <Description>\nPDGM Group: <PDGM Group>\nComorbidity Group: <Comorbidity Group>\n"
        "Primary Awarding: <Primary Awarding Flag>\nCode First: <Code First>\nBillable: <Yes/No (Section 111 Excluded)>\n"
        "Reason: <Brief explanation for code selection>\nSeparate multiple codes with a blank line. Never invent codes."
    )
    candidates = get_candidate_codes(phrase)
    code_choices = "\n".join(
        f"{row['icd10_code']}|{row.get('description','')}|{row.get('pdgm_clinical_group_code','')}|"
        f"{row.get('pdgm_clinical_group_name','')}|{row.get('COMORBIDITY_GROUP','')}|"
        f"{row.get('PRIMARY_AWARDING_FLAG','')}|{row.get('CODE_FIRST','')}"
        for row in candidates
    )
    exclusions = ", ".join(list(excluded_codes)[:30]) + ("..." if len(excluded_codes) > 30 else "")
    user_prompt = (
        f"Clinical input: '{phrase}'\n"
        "Codes list (ICD10|Description|PDGM Code|PDGM Group|Comorbidity Group|Primary Awarding|Code First):\n"
        f"{code_choices}\n"
        f"Codes excluded as primary: {exclusions}\n"
        "Return the closest code(s) in the specified format."
    )

    try:
        text = _call_claude(system_prompt, user_prompt, max_tokens=350)
        valid, mismatches = validate_ai_output(text)
        log_ai_interaction(CLAUDE_MODEL, phrase, text, mismatches)
        return text
    except anthropic.RateLimitError:
        return "Rate limit exceeded. Please try again later."
    except anthropic.APIError as e:
        return f"AI error: {e}"


def ai_followup_question(phrase: str) -> str:
    fallback_response = get_fallback_followup(phrase)
    if fallback_response != "Could you provide more specific details about the condition, such as: acute vs chronic, location, severity, or any complications?":
        return fallback_response
    if not client:
        return fallback_response
    try:
        return _call_claude(
            "You are a medical coding specialist helping refine ICD-10 selection for home health PDGM coding. ALWAYS provide helpful follow-up questions with specific options.",
            f"Medical term: '{phrase}' - What follow-up questions would help specify this for accurate ICD-10 coding?",
            max_tokens=120,
            temperature=0.7,
        )
    except Exception:
        pass
    return fallback_response


# ---------------------------------------------------------------------------
# Documentation roadmap & OASIS assessment
# ---------------------------------------------------------------------------

def extract_pdgm_info_from_response(pdgm_response):
    lines = pdgm_response.split('\n')
    diagnosis = pdgm_group = None
    for line in lines:
        line = line.strip()
        if line.startswith('Focus of Care:'):
            parts = line.split()
            for part in parts:
                if '.' in part and len(part) >= 4:
                    diagnosis = part
                    break
        elif line.startswith('PDGM Group:'):
            pdgm_group = line.replace('PDGM Group:', '').strip()
    return diagnosis, pdgm_group


def ai_documentation_roadmap(diagnosis: str, pdgm_group: str, disciplines: list):
    if not client:
        return "Anthropic API key not configured."
    try:
        system_prompt = prompt_manager.build_system_prompt("roadmap", pdgm_group, diagnosis, disciplines)
        user_prompt = f"Diagnosis: {diagnosis}\nPDGM Group: {pdgm_group}\nDisciplines: {', '.join(disciplines)}"
        return _call_claude(system_prompt, user_prompt, max_tokens=1000)
    except Exception as e:
        return f"AI error: {e}"


def ai_sample_oasis_assessment(diagnosis: str, pdgm_group: str, disciplines: list = None):
    if not client:
        return "Anthropic API key not configured."
    try:
        system_prompt = prompt_manager.build_system_prompt("oasis", pdgm_group, diagnosis, disciplines or [])
        user_prompt = f"Diagnosis: {diagnosis}\nPDGM Group: {pdgm_group}"
        return _call_claude(system_prompt, user_prompt, max_tokens=1000)
    except Exception as e:
        return f"AI error: {e}"
